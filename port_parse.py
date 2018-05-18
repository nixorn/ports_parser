#!/usr/bin/python3
import argparse
import openpyxl
import re

from config import sheet_structure
from grammar import parser
from postprocessing import bring_beauty


PARSER = argparse.ArgumentParser(description='Port situation xls extraction tool.')


PARSER.add_argument('--file', nargs='?',
                    help='Path to source file.')

if __name__ == '__main__':
    ARGS = PARSER.parse_args()
    if not ARGS.file:
        PARSER.print_help()
    print(
        "ReportDate;" +\
        "Port;" +\
        "Vesselstatus;" +\
        "Pier;" +\
        "Vesselname;" +\
        "ETA;" +\
        "ETS;" +\
        "Grade;" +\
        "Quantity;" +\
        "LastPort;" +\
        "NextPort;" +\
        "Sailed;" +\
        "ETB")
    wb = openpyxl.load_workbook(ARGS.file)
    parsed_sheets = []
    for sheet, sheet_name in zip(wb, wb.sheetnames):
        matrix = [[cell.value for cell in row] for row in sheet.iter_rows()]
        parsed_sheets += parser(sheet_structure, matrix, sheet_name, ARGS.file)
    for row in bring_beauty(parsed_sheets):
        print(";".join(row))
